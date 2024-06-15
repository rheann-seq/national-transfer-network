import openpyxl

wb = openpyxl.load_workbook('./files/WVU Transfer Credit Database.xlsx')

ws = wb.active

print('Total number of rows: '+str(ws.max_row)+'. Total number of columns: '+str(ws.max_column))

